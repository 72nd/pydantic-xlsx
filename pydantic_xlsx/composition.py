"""
pydantic-xlsx supports different kind of table/model compositions/structures.
This can be one of the following cases:

- Single Model Composition: The given pydantic schema doesn't contain any list
  of other Models. In this case the data will be exported into a single Sheet.
  The sheet will be named after the schema's title or (if not present) the
  class name of the model.
- Root Collection Composition: The Model consists of a single __root__ property
  which is a list with instances of a model. The resulting document will
  contain a single sheet named after the title of the model. If the title is
  not present, the name of the model class (containing the __root__ property)
  will be used instead.
- Collection Composition: The model consists of at least one list of model
  instances. There will be a sheet for each property (even for the properties
  which are not lists).
"""


from abc import ABCMeta, abstractmethod
from enum import Enum
import math
from typing import Any, Dict, List, Type, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl import utils as pyxl_utils
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

from .config import XlsxConfig
from .fields import FieldTypeInfoFactory

if TYPE_CHECKING:
    from .model import XlsxModel

EXCEL_MAX_ROWS = 1048576


class Composition(metaclass=ABCMeta):
    """
    Base class for all supported composition cases. Learn more about this cases
    in the module documentation. The composition should be initialized via the
    `CompositionFactory`. This class also handles much of the heavy-lifting for
    the actual conversion between Excel Sheets and XlsxModels.
    """

    data: List["XlsxModel"]
    title: str
    config: XlsxConfig

    def __init__(
            self,
            data: List["XlsxModel"],
            title: str,
            config: XlsxConfig
    ):
        if len(data) < 1:
            raise ValueError("cannot create the {} sheet, at least one entry \
is needed. This error also can occurs when using aliases for model \
collections without setting allow_population_by_field_name to \
True.".format(title))
        for model in data:
            if len(model.__fields__) == 0:
                raise ValueError(
                    "Model {} hasn't got any properties".format(
                        model.__class__.__name__)
                )
        self.data = data
        self.title = title
        self.config = config

    @classmethod
    @abstractmethod
    def sheets_from_model(cls, model: "XlsxModel") -> List["Composition"]:
        """
        Takes a `XlsxModel` and rearranges the data to exportable sheets.
        """
        pass

    @classmethod
    @abstractmethod
    def workbook_to_model(
        cls,
        model: Type["XlsxModel"],
        wb: Workbook,
    ) -> "XlsxModel":
        """
        Takes a openpyxl Workbook and parses it to a `XlsxModel`.
        """
        pass

    def populate_worksheet(self, ws: Worksheet):
        """
        Populates a given Excel worksheet with data and title.
        """
        ws.title = self.title
        self.__worksheet_insert_header(ws)
        self.__worksheet_insert_content(ws)
        self.__freeze_cells(ws)
        self.__apply_apperance(ws)
        self.__apply_print_settings(ws)
        self.__data_table(ws)

    def __worksheet_insert_header(self, ws: Worksheet) -> None:
        """Sets and formats the header of the Worksheet."""
        header_row = ws.row_dimensions[1]
        if (font := self.config.header_font) is not None:
            header_row.font = font
        if (alignment := self.config.header_alignment) is not None:
            header_row.alignment = alignment
        ws.append(self.data[0]._property_keys())

    def __worksheet_insert_content(self, ws: Worksheet) -> None:
        """Inserts the actual data into the Worksheet."""
        for row in self.data:
            ws.append([cell for cell in row.dict().values()])

    def __freeze_cells(self, ws: Worksheet) -> None:
        """Freezes the cell at the given cell from the config."""
        if (freeze_at := self.config.freeze_cell) is not None:
            ws.freeze_panes = ws[freeze_at]

    def __apply_apperance(self, ws: Worksheet) -> None:
        """
        Apply column-specific formatting. Formatting settings from specific
        rows (as defined by `.XlsxField`) will be preferred over document wide
        settings. Also some Field Types (like `.types.Money`) which have a
        `.fields.FieldTypeInfo` implemented can derive `.fields.XlsxFieldInfo`.
        Also in this case they will be overturned user specific configurations
        by a `.XlsxField` in the model.
        """
        fields = list(self.data[0].__fields__.values())
        lengths = self.__calc_column_widths()

        # TODO: Refactoring this beast.

        for col in range(1, ws.max_column+1):
            column_letter = pyxl_utils.cell.get_column_letter(col)
            column = ws.column_dimensions[column_letter]
            field_info = fields[col-1].field_info
            field_type_info = FieldTypeInfoFactory.field_info_from_type(
                fields[col-1].type_
            )

            # Apply List DataValidation to column.
            if issubclass(fields[col-1].type_, Enum):
                dv = self.__validator_for_enum(fields[col-1].type_)
                ws.add_data_validation(dv)
                dv.add(f"{column_letter}1:{column_letter}{EXCEL_MAX_ROWS}")

            # 1st font priority: XlsxField.
            if hasattr(field_info, "font") and \
                    field_info.font is not None:
                column.font = field_info.font
            # 2nd font priority: Derived from FieldTypeInfo.
            elif hasattr(field_type_info, "font") and \
                    field_type_info.font is not None:
                column.font = field_info.font
            # 3rd font priority: Document wide setting.
            elif self.config.font is not None:
                column.font = self.config.font

            # 1st number format priority: XlsxField.
            if hasattr(field_info, "number_format") and \
                    field_info.number_format is not None:
                column.number_format = field_info.number_format
            # 2nd number format priority: Derived from FieldTypeInfo.
            elif hasattr(field_type_info, "number_format") and \
                    field_type_info.number_format is not None:
                column.number_format = field_type_info.number_format

            column.width = lengths[col-1]

    def __apply_print_settings(self, ws: Worksheet) -> None:
        """Applies the print settings."""
        ws.print_options.horizontalCentered =\
            self.config.print_horizontal_centered
        ws.print_options.verticalCentered =\
            self.config.print_vertical_centered
        ws.print_title_cols = self.config._print_title_columns()
        ws.print_title_rows = self.config._print_title_rows()

    def __data_table(self, ws: Worksheet) -> None:
        """Configures the data table."""
        table = Table(
            displayName=self.title.replace(" ", ""),
            ref=self.__dimensions()
        )
        ws.add_table(table)

    def __dimensions(self) -> str:
        """Returns the dimension of the used space."""
        return "A1:{col}{row}".format(
            col=pyxl_utils.cell.get_column_letter(
                len(self.data[0].__fields__)),
            row=len(self.data) + 1,
        )

    def __calc_column_widths(self) -> List[int]:
        """Calculates an approximative width for each column."""
        widths_per_key: Dict[str, int] = {}
        keys = [key for key in self.data[0].__fields__]
        for key in keys:
            widths_per_key[key] = 0

        for entry in self.data:
            entry_dict = entry.dict()
            for key in keys:
                if not isinstance(entry_dict[key], str):
                    continue
                if (length := len(entry_dict[key])) > widths_per_key[key]:
                    widths_per_key[key] = length
        rsl: List[int] = []
        for value in widths_per_key.values():
            if (length := math.ceil(value*0.93)) > 5:
                rsl.append(length)
            else:
                rsl.append(5)
        return rsl

    @staticmethod
    def __validator_for_enum(enum: Enum) -> DataValidation:
        """Returns a DataValidation for a given enum."""
        formula = '"{}"'.format(",".join([item.value for item in enum]))
        name = enum.__name__

        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Your entry is not a valid {}".format(name)
        dv.errorTitle = "Invalid {}".format(name)
        dv.prompt = "Select {} from the list".format(name)
        dv.promptTitle = "Select {}".format(name)

        return dv

    def __str__(self) -> str:
        return "{} ({})".format(self.title, self.data)

    def __repr__(self):
        return self.__str__()


class SingleModelComposition(Composition):
    """
    The Model doesn't contain any lists of other Models thus will be exported
    in a single Sheet (named after the Schema title or if this is not present
    the Model class name.
    """

    @classmethod
    def sheets_from_model(
        cls,
        model: "XlsxModel"
    ) -> List["SingleModelComposition"]:
        if (schema_title := model.__config__.title) is not None:
            title = schema_title
        else:
            title = model.__class__.__name__
        return [cls([model], title, model.__config__)]

    @classmethod
    def workbook_to_model(
        cls,
        model: Type["XlsxModel"],
        wb: Workbook,
    ) -> "XlsxModel":
        ws = wb.active
        entry: Dict[str, Any] = {}
        for column in range(1, ws.max_column+1):
            key = ws.cell(row=1, column=column).value
            if model._ignore_key(key):
                continue
            entry[key] = ws.cell(row=2, column=column).value
        return model.parse_obj(entry)


class RootCollectionComposition(Composition):
    """
    The Model consists of a single __root__ property which is a list of a
    Model. The document will contain a single sheet named after the title or
    if this is not present the name of the Model class the collection contains.
    """

    @classmethod
    def sheets_from_model(
        cls,
        model: "XlsxModel"
    ) -> List["RootCollectionComposition"]:
        if (schema_title := model.__config__.title) is not None:
            title = schema_title
        else:
            title = model.__root__[0].__class__.__name__
        return [cls(model.__root__, title, model.__config__)]

    @classmethod
    def workbook_to_model(
        cls,
        model: Type["XlsxModel"],
        wb: Workbook,
    ) -> "XlsxModel":
        ws = wb.active

        data: List[Dict[str, Any]] = []
        for row in range(2, ws.max_row + 1):
            entry: Dict[str, Any] = {}
            for column in range(1, ws.max_column+1):
                key = ws.cell(row=1, column=column).value
                if model._ignore_key(key):
                    continue
                entry[key] = ws.cell(row=row, column=column).value
            data.append(entry)
        return model.parse_obj(data)


class CollectionComposition(Composition):
    """
    The model consists of at least one list of XlsxModel instances. There will
    be a sheet for each property.
    """

    @classmethod
    def sheets_from_model(
            cls, model: "XlsxModel"
    ) -> List["CollectionComposition"]:
        rsl: List["CollectionComposition"] = []
        for key, field in model.__fields__.items():
            rsl.append(cls(
                model.__dict__[key],
                field.alias,
                model.__config__
            ))
        return rsl

    @classmethod
    def workbook_to_model(
        cls,
        model: Type["XlsxModel"],
        wb: Workbook,
    ) -> "XlsxModel":
        data: Dict[str, List[Dict[str, Any]]] = {}

        for prop in model.__fields__.values():
            ws = wb[prop.alias]
            data[prop.alias] = []
            current_model = model.__fields__[prop.alias].type_
            for row in range(2, ws.max_row+1):
                entry: Dict[str, Any] = {}
                for column in range(1, ws.max_column+1):
                    key = ws.cell(row=1, column=column).value
                    if current_model._ignore_key(key):
                        continue
                    entry[key] = ws.cell(row=row, column=column).value
                data[prop.alias].append(entry)
        return model.parse_obj(data)


class CompositionFactory:
    """
    Creates the correct `Composition` implementation based on the structure of
    a given model.
    """

    @classmethod
    def from_model(cls, model: "XlsxModel") -> Composition:
        """
        Creates and returns the correct `Composition` implementation for the
        given Model.
        """
        contains_list_of_models = False

        for key, prop in model.__fields__.items():
            typ = getattr(prop.outer_type_, "__origin__", None)

            # Thanks to Python's inability to handle cyclic imports checking if
            # the type is XlsxModel is done with this handy string hack.
            bases = [base.__name__ for base in prop.type_.__bases__]
            if typ is not None and \
                    issubclass(typ, List) and \
                    "XlsxModel" in bases:
                contains_list_of_models = True
            elif typ is not None and issubclass(typ, Dict):
                raise TypeError("Dicts currently not supported")

        # If __root__ is present as a parameter in a model pydantic forbids any
        # other fields. Thus it's not needed to check if there are any other
        # lists of models.
        if "__root__" in model.__fields__ and contains_list_of_models:
            return RootCollectionComposition
        if contains_list_of_models:
            return CollectionComposition
        return SingleModelComposition
