"""
Provides the `XlsxModel` to the library user.
"""

from pathlib import Path
from typing import Any, List, Optional, Type, Union

from openpyxl import load_workbook, Workbook
from pydantic import BaseModel

from .composition import CompositionFactory
from .config import XlsxConfig


class XlsxModel(BaseModel):
    """
    Extends pydantic with the ability to in- and export data from/to xlsx
    files. The class also allows the usage of collection by using the __root__
    parameter.
    """

    Config = XlsxConfig

    def __init__(self, **data: Any):
        super().__init__(**data)

    @classmethod
    def from_file(
            cls: Type["XlsxModel"],
            path: Union[str, Path]
    ) -> "XlsxModel":
        """Loads data from a Excel file."""
        return cls.from_workbook(load_workbook(path))

    @classmethod
    def from_workbook(
            cls: Type["XlsxModel"],
            wb: Workbook
    ) -> "XlsxModel":
        """Loads data from a Excel file."""

        return CompositionFactory.from_model(cls).workbook_to_model(cls, wb)

    def workbook(self) -> Workbook:
        """Returns a openpyxl Workbook."""
        wb = Workbook()
        sheets = CompositionFactory.from_model(self).sheets_from_model(self)

        first_sheet = True
        for sheet in sheets:
            if first_sheet:
                ws = wb.active
                first_sheet = False
            else:
                ws = wb.create_sheet()
            sheet.populate_worksheet(ws)
        return wb

    def to_file(self, path: Union[str, Path]):
        """Saves the model to a Excel file."""
        wb = self.workbook()
        wb.save(path)

    @classmethod
    def _property_keys(cls) -> List[str]:
        """Returns the names of the properties used in the excel header."""
        return [field.alias for field in cls.__fields__.values()]

    @classmethod
    def _ignore_key(cls, key: Optional[str]) -> bool:
        """
        Takes a property name (key) and checks if ignore_additional_rows is
        enabled the property is described by the model. This method is used by
        the from_workbook methods to determine if a cell should be added to the
        import dict.
        """
        ignore = getattr(cls.__config__, "ignore_additional_rows", None)
        if ignore is not None and not ignore:
            return False
        if key is None:
            return True
        return key not in cls._property_keys()
